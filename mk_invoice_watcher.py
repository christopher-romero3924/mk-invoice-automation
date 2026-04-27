# MK Invoice Watcher
# ------------------
# Watches the Working Invoices folder for completed invoices.
# When an invoice has E36 = "YES" and the file is closed (not locked),
# it parses the invoice, writes to Raw_Invoice_Data and the correct
# year sheet in Priority_use, then moves the file to Complete Invoices.
#
# Setup:
#   1. Install dependencies:  pip install openpyxl watchdog
#   2. Run once manually to test:  python mk_invoice_watcher.py
#   3. To run on startup, add to Windows Task Scheduler (see README at bottom)
#
# Folder structure expected:
#   C:/Users/Christopher/Desktop/MK Automation Test/
#       Working Invoices/        <-- drop invoices here
#       Complete Invoices/       <-- processed invoices move here
#       Priority_use_2026_CLEAN_3.xlsx

import os
import re
import time
import shutil
import logging
import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ── CONFIG ────────────────────────────────────────────────────────────────────
BASE_DIR        = r"C:\Users\Christopher\Desktop\MK Automation Test"
WORKING_DIR     = os.path.join(BASE_DIR, "Working Invoices")
COMPLETE_DIR    = os.path.join(BASE_DIR, "Complete Invoices")
PRIORITY_FILE   = os.path.join(BASE_DIR, "Priority_use_2026_CLEAN_3.xlsx")
LOG_FILE        = os.path.join(BASE_DIR, "mk_watcher.log")
CHECK_DELAY     = 3   # seconds to wait after file change before processing

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ── COLUMN MAP ────────────────────────────────────────────────────────────────
PRIORITY_COL = {1: 2, 2: 3, 3: 4}   # Priority level -> sheet column (B/C/D)

RAW_COLS = [
    'Invoice_Number', 'Group_Name', 'Start_Date', 'End_Date',
    'Priority_Level', 'Number_of_Campers', 'Days', 'Camper_Days',
    'Year', 'Contact_Name', 'Phone', 'Explanation'
]


def is_file_locked(path):
    """Check if Excel still has the file open (locked)."""
    try:
        os.rename(path, path)
        return False
    except OSError:
        return True


def is_completed(path):
    """Return True if E36 == 'YES' (case-insensitive)."""
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        val = ws['E36'].value
        return str(val).strip().upper() == 'YES' if val else False
    except Exception as e:
        log.warning(f"Could not read {path}: {e}")
        return False


def parse_invoice(path):
    """Parse invoice fields and return a dict."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    data = {}

    for row in ws.iter_rows(min_row=1, max_row=35):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()

            if cell.column == 2 and cell.row == 6:
                data['Group_Name'] = s
            if cell.column == 5 and cell.row == 6:
                data['Invoice_Number'] = v
            if 'DATE START:' in s:
                m = re.search(r'(\w+ \d+, \d{4})', s)
                if m:
                    data['Start_Date'] = datetime.datetime.strptime(m.group(1), '%B %d, %Y')
            if 'DATE END:' in s:
                m = re.search(r'(\w+ \d+, \d{4})', s)
                if m:
                    data['End_Date'] = datetime.datetime.strptime(m.group(1), '%B %d, %Y')
            if s.startswith('Contact:'):
                data['Contact_Name'] = s.replace('Contact:', '').strip()
            if s.startswith('Phone:'):
                data['Phone'] = s.replace('Phone:', '').strip()
            if s.startswith('Campers:'):
                m = re.search(r'\d+', s)
                if m: data['Number_of_Campers'] = int(m.group())
            if s.startswith('Priority:'):
                m = re.search(r'\d+', s)
                if m: data['Priority_Level'] = int(m.group())
            if s.startswith('Days:'):
                m = re.search(r'\d+', s)
                if m: data['Days'] = int(m.group())
            if s.startswith('Explanation:'):
                data['Explanation'] = s.replace('Explanation:', '').strip()

    if 'Start_Date' in data:
        data['Year'] = data['Start_Date'].year
    if 'Number_of_Campers' in data and 'Days' in data:
        data['Camper_Days'] = data['Number_of_Campers'] * data['Days']

    return data


def get_next_empty_row(ws, start=2):
    """Find the next empty row in Raw_Invoice_Data."""
    r = start
    while True:
        if all(ws.cell(row=r, column=c).value is None for c in range(1, 13)):
            return r
        r += 1


def update_table_ref(ws, last_row):
    """Resize the InvoiceData table to cover new last row."""
    if 'InvoiceData' in ws.tables:
        del ws.tables['InvoiceData']
    tbl = Table(displayName="InvoiceData", ref=f"A1:L{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)


def get_year_sheet(wb, year):
    """Return the sheet matching the invoice year (e.g. '26' for 2026)."""
    short = str(year)[-2:]   # 2026 -> '26'
    if short in wb.sheetnames:
        return wb[short]
    log.warning(f"No sheet named '{short}' found for year {year}. Available: {wb.sheetnames}")
    return None


def get_next_empty_row_sheet(ws, start=6, col=1):
    """Find next empty row in a year sheet (data starts at row 6)."""
    r = start
    while ws.cell(row=r, column=col).value is not None:
        r += 1
        if r > 200:
            break
    return r


def post_invoice(inv, priority_file):
    """Write parsed invoice to Raw_Invoice_Data and the year sheet."""
    wb = load_workbook(priority_file)
    ws_raw = wb['Raw_Invoice_Data']

    # ── Write to Raw_Invoice_Data ─────────────────────────────────────────────
    next_row = get_next_empty_row(ws_raw)
    for c_idx, col in enumerate(RAW_COLS, start=1):
        ws_raw.cell(row=next_row, column=c_idx).value = inv.get(col)
    update_table_ref(ws_raw, next_row)
    log.info(f"  Raw_Invoice_Data row {next_row}: {inv.get('Group_Name')}")

    # ── Write to year sheet ───────────────────────────────────────────────────
    year = inv.get('Year')
    ws_year = get_year_sheet(wb, year)

    if ws_year:
        target_row = get_next_empty_row_sheet(ws_year)
        priority   = inv.get('Priority_Level')
        p_col      = PRIORITY_COL.get(priority)

        ws_year.cell(row=target_row, column=1).value  = inv.get('Group_Name')
        if p_col:
            ws_year.cell(row=target_row, column=p_col).value = inv.get('Number_of_Campers')
        ws_year.cell(row=target_row, column=5).value  = inv.get('Days')

        start_cell = ws_year.cell(row=target_row, column=9)
        end_cell   = ws_year.cell(row=target_row, column=10)
        start_cell.value        = inv.get('Start_Date')
        start_cell.number_format = 'm/d/yyyy'
        end_cell.value          = inv.get('End_Date')
        end_cell.number_format  = 'm/d/yyyy'

        ws_year.cell(row=target_row, column=11).value = inv.get('Contact_Name')
        ws_year.cell(row=target_row, column=12).value = inv.get('Phone')
        ws_year.cell(row=target_row, column=13).value = inv.get('Explanation')

        log.info(f"  Sheet '{ws_year.title}' row {target_row}: {inv.get('Group_Name')} | P{priority} | {inv.get('Number_of_Campers')} campers")

    wb.save(priority_file)


def process_file(path):
    """Full processing pipeline for one invoice file."""
    filename = os.path.basename(path)
    log.info(f"Checking: {filename}")

    # Wait for Excel to release the file
    for _ in range(10):
        if not is_file_locked(path):
            break
        log.info("  File still locked, waiting...")
        time.sleep(2)
    else:
        log.warning(f"  File still locked after retries, skipping: {filename}")
        return

    if not is_completed(path):
        log.info(f"  Not marked complete, skipping.")
        return

    log.info(f"  Marked COMPLETE — parsing...")
    inv = parse_invoice(path)

    required = ['Group_Name', 'Start_Date', 'End_Date', 'Priority_Level',
                'Number_of_Campers', 'Days', 'Year']
    missing = [f for f in required if f not in inv]
    if missing:
        log.error(f"  Missing fields: {missing} — skipping {filename}")
        return

    log.info(f"  Parsed: {inv.get('Group_Name')} | P{inv.get('Priority_Level')} | "
             f"{inv.get('Number_of_Campers')} campers x {inv.get('Days')} days | "
             f"Year {inv.get('Year')}")

    post_invoice(inv, PRIORITY_FILE)

    # Move to Complete Invoices
    dest = os.path.join(COMPLETE_DIR, filename)
    if os.path.exists(dest):
        base, ext = os.path.splitext(filename)
        dest = os.path.join(COMPLETE_DIR, f"{base}_{int(time.time())}{ext}")
    shutil.move(path, dest)
    log.info(f"  Moved to Complete Invoices: {os.path.basename(dest)}")


# ── FILE WATCHER ──────────────────────────────────────────────────────────────
class InvoiceHandler(FileSystemEventHandler):
    def __init__(self):
        self.pending = {}   # path -> timestamp of last change

    def on_modified(self, event):
        if event.is_directory:
            return
        path = event.src_path
        if not path.lower().endswith('.xlsx'):
            return
        if os.path.basename(path).startswith('~$'):
            return   # skip Excel temp files
        self.pending[path] = time.time()

    def on_created(self, event):
        self.on_modified(event)

    def flush_pending(self):
        """Process files that haven't changed in CHECK_DELAY seconds."""
        now = time.time()
        ready = [p for p, t in list(self.pending.items()) if now - t >= CHECK_DELAY]
        for path in ready:
            del self.pending[path]
            if os.path.exists(path):
                try:
                    process_file(path)
                except Exception as e:
                    log.error(f"Error processing {path}: {e}", exc_info=True)


def main():
    log.info("=" * 60)
    log.info("MK Invoice Watcher started")
    log.info(f"  Watching:       {WORKING_DIR}")
    log.info(f"  Complete folder:{COMPLETE_DIR}")
    log.info(f"  Priority file:  {PRIORITY_FILE}")
    log.info("=" * 60)

    os.makedirs(WORKING_DIR,  exist_ok=True)
    os.makedirs(COMPLETE_DIR, exist_ok=True)

    handler  = InvoiceHandler()
    observer = Observer()
    observer.schedule(handler, WORKING_DIR, recursive=False)
    observer.start()

    try:
        while True:
            handler.flush_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        log.info("Watcher stopped by user.")
        observer.stop()
    observer.join()


if __name__ == '__main__':
    main()


# ══════════════════════════════════════════════════════════════════════════════
# WINDOWS TASK SCHEDULER SETUP (run once to register)
# ══════════════════════════════════════════════════════════════════════════════
#
# 1. Install Python if not already installed: https://python.org
#    Make sure to check "Add Python to PATH" during install.
#
# 2. Open Command Prompt as Administrator and run:
#       pip install openpyxl watchdog
#
# 3. Save this script to:
#       C:/Users/Christopher/Desktop/MK Automation Test/mk_invoice_watcher.py
#
# 4. Open Task Scheduler (search "Task Scheduler" in Start menu)
#
# 5. Click "Create Basic Task" and fill in:
#       Name:        MK Invoice Watcher
#       Trigger:     When I log on
#       Action:      Start a program
#       Program:     pythonw.exe          (runs silently, no console window)
#       Arguments:   "C:/Users/Christopher/Desktop/MK Automation Test/mk_invoice_watcher.py"
#
# 6. Click Finish. The watcher will now start automatically when you log in.
#
# 7. To check if it's working, look at mk_watcher.log in the MK Automation Test folder.
#
# ══════════════════════════════════════════════════════════════════════════════
